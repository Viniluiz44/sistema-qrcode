from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
from datetime import date
import os

app = Flask(__name__)
arquivo_excel = "manutencoes.xlsx"

class SistemaManutencao:
    def __init__(self, arquivo=arquivo_excel):
        self.arquivo = arquivo

    def criar_arquivo_excel(self):
        wb = Workbook()

        # Planilha de Equipamentos
        ws1 = wb.active
        ws1.title = "Equipamentos"
        ws1.append(["ID", "Nome", "QRCode"])

        # Planilha de Manutencoes Preventivas
        ws2 = wb.create_sheet(title="Manutencoes")
        ws2.append(["Equipamento_ID", "Data_Manutencao", "Descricao"])

        wb.save(self.arquivo)

    def cadastrar_equipamento(self, nome_equipamento, qr_code):
        wb = load_workbook(self.arquivo)
        ws = wb["Equipamentos"]
        next_id = ws.max_row  # ID é baseado na quantidade de linhas já preenchidas
        ws.append([next_id, nome_equipamento, qr_code])
        wb.save(self.arquivo)

    def registrar_manutencao(self, qr_code, descricao):
        wb = load_workbook(self.arquivo)
        ws_equip = wb["Equipamentos"]

        equipamento_id = None
        for row in ws_equip.iter_rows(min_row=2, values_only=True):
            if row[2] == qr_code:
                equipamento_id = row[0]
                break

        if equipamento_id is None:
            return False

        ws_manut = wb["Manutencoes"]
        ws_manut.append([equipamento_id, str(date.today()), descricao])
        wb.save(self.arquivo)
        return True

    def consultar_historico_por_qr(self, qr_code):
        wb = load_workbook(self.arquivo)

        # Buscar o equipamento pelo QR Code
        ws_equip = wb["Equipamentos"]
        equipamento_id = None
        nome_equipamento = None
        for row in ws_equip.iter_rows(min_row=2, values_only=True):
            if row[2] == qr_code:
                equipamento_id = row[0]
                nome_equipamento = row[1]
                break

        if equipamento_id is None:
            return None

        # Buscar manutenções do equipamento
        ws_manut = wb["Manutencoes"]
        manutencoes = []
        for row in ws_manut.iter_rows(min_row=2, values_only=True):
            if row[0] == equipamento_id:
                manutencoes.append({"data": row[1], "descricao": row[2]})

        return {"nome": nome_equipamento, "manutencoes": manutencoes}


sistema = SistemaManutencao()

# Rota principal
@app.route('/')
def index():
    return render_template('index.html')

# Rota para criar o arquivo Excel
@app.route('/criar_arquivo')
def criar_arquivo():
    sistema.criar_arquivo_excel()
    return "Arquivo Excel criado com sucesso! <br><a href='/'>Voltar</a>"

# Rota para cadastrar um novo equipamento
@app.route('/cadastrar_equipamento', methods=['GET', 'POST'])
def cadastrar_equipamento():
    if request.method == 'POST':
        nome_equipamento = request.form['nome_equipamento']
        qr_code = request.form['qr_code']
        sistema.cadastrar_equipamento(nome_equipamento, qr_code)
        return redirect(url_for('index'))
    return render_template('cadastrar_equipamento.html')

# Rota para registrar uma manutenção
@app.route('/registrar_manutencao', methods=['GET', 'POST'])
def registrar_manutencao():
    if request.method == 'POST':
        qr_code = request.form['qr_code']
        descricao = request.form['descricao']
        sucesso = sistema.registrar_manutencao(qr_code, descricao)
        if sucesso:
            return "Manutenção registrada com sucesso! <br><a href='/'>Voltar</a>"
        else:
            return "QR Code não encontrado! <br><a href='/registrar_manutencao'>Tente novamente</a>"
    return render_template('registrar_manutencao.html')

# Rota para consultar o histórico por QR Code
@app.route('/consultar_historico', methods=['GET', 'POST'])
def consultar_historico():
    if request.method == 'POST':
        qr_code = request.form['qr_code']
        historico = sistema.consultar_historico_por_qr(qr_code)
        if historico:
            return render_template('historico.html', historico=historico)
        else:
            return "QR Code não encontrado! <br><a href='/consultar_historico'>Tente novamente</a>"
    return render_template('consultar_historico.html')


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))  # Usa a porta do ambiente ou 5000
    app.run(host='0.0.0.0', port=port)